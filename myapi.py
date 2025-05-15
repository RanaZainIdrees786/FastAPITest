from openpyxl import Workbook, load_workbook
import os
from groq import Groq
import pyttsx3 
import speech_recognition as sr
from emailSending import send_email
import pywhatkit as kit
import time
from whatsappSend import send_measage
from readData import read_excel_data
from readData import read_pdf_data

client = Groq(
    api_key="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",
)

# initialiezed text to speech engine
engine = pyttsx3.init()
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[37].id)


def recognize_speech():
# Initialize recognizer
    recognizer = sr.Recognizer()      
        # Use microphone as input
    with sr.Microphone() as source:
        print("🎤 Speak something...")
        recognizer.adjust_for_ambient_noise(source)  # Optional: better accuracy
        audio = recognizer.listen(source)
        try:
            # Recognize speech using Google Speech Recognition
            text = recognizer.recognize_google(audio)
            return text
        except sr.UnknownValueError:
            print("❌ Sorry, I could not understand your speech.")
            return None
        except sr.RequestError:
            print("🚫 Speech Recognition service is unavailable.")
            return None

IsContinue = True
while IsContinue:
    # user_question = recognize_speech()
    # if user_question == None:
    #     continue
    # engine.say(user_question)
    # engine.runAndWait()
    akti_info = read_pdf_data('data/Arfa Karim Technology Incubator.pdf')
    courses_info = read_excel_data('data/Courses.xlsx')
    user_question = input("Ask a question about AKTI or type exit: ")


    prompt_template = f"""
            Rules:
            1. The provided answer should not contain more than two lines
            2. If someone asks you about course dont disclose information about teacher
            3. Answers should only include what has been asked
            4. provide answers not more than two line?
            5. provide answers only from the given context.


            Context:

                Courses: 
                {courses_info}

                AKTI INFO:
                {akti_info}

            User Question:    
            {user_question}            
    """

    chat_completion = client.chat.completions.create(
        messages=[
            {
                "role": "user",
                "content": prompt_template,
            }
        ],
        model="llama-3.3-70b-versatile",
    )

    response  = chat_completion.choices[0].message.content
    print(response)
    engine.say(response)
    engine.runAndWait()


    enrollment = input("Would you like to enroll in any of the courses? y for yes n for no: \n")
    excel_file = "akti_enrollments.xlsx"
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Contact", "Email", "Course", "Batch", "time"])
        wb.save(excel_file)


    match enrollment:
        case "y":
            
            name = input("please enter your name:\n")
            contact = input("Tell us your contact number on watsapp:\n")
            email = input("Please enter your email:\n")
            course = input("In which course would you like to enroll:\n")
            meetup_time = input("Your fee will be $400 with discount \n you will have to come for an interview with the manager \n on what time would you like to come ?: ")
            batch = input("Which batch would you like to enroll your self: \n 11 \n 12\n 13\n")

            match batch:
                case "11":
                    print("Your Batch No. is: 11")
                case "12":
                    print("Your Batch No. is: 12")
                case "13":
                    print("Your Batch No. is: 13")

            wb = load_workbook(excel_file)
            ws = wb.active
            ws.append([name, contact, email, course, batch, meetup_time])
            wb.save(excel_file) 
            email_text =  f"Welcome {name} to AKTI! You have enrolled yourself in {course} and your batch is {batch}. Please come on {time}.Tomorrow we will send you your nearest location on watsapp!"
            send_email(email,email_text)
            send_measage(contact,email_text)
            # print(email_text)
            ask = input("would you like to ask more y for yes n for no: ")
            if ask == "n":
                    IsContinue = False
        case "n":
                ask = input("would you like to ask more y for yes n for no: ")
                match ask:
                    case "y":
                        pass
                    case "n":
                        IsContinue = False

    # email = chat_completion.choices[0].message.content    import pyttsx3


